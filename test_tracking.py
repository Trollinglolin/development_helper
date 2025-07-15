import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from tabulate import tabulate

class TestCaseTracker:
    def __init__(self, filename="test_cases.xlsx"):
        self.filename = filename
        self.headers = ["Index", "Objective", "Date", "Person", "Expectation", "Results", "Remark", "Status"]
        try:
            self.wb = openpyxl.load_workbook(filename)
            self.ws = self.wb.active
            # Verify headers exist in the sheet
            if not all(self.ws.cell(row=1, column=i+1).value == self.headers[i] for i in range(len(self.headers))):
                self._initialize_workbook()
        except FileNotFoundError:
            self._initialize_workbook()
    
    def _initialize_workbook(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Test Cases"
        for col_num, header in enumerate(self.headers, 1):
            cell = self.ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        self.wb.save(self.filename)
    
    def _get_next_index(self):
        return self.ws.max_row
    
    def add_test_case(self, objective, person, expectation):
        index = self._get_next_index()
        self.ws.append([
            index,
            objective,
            datetime.now().strftime("%Y-%m-%d"),
            person,
            expectation,
            "",  # Results
            "",  # Remark
            "Pending"  # Status
        ])
        self.wb.save(self.filename)
        return index
    
    def update_test_case(self, index, objective, date, person, expectation, results, remark, status):
        for row in self.ws.iter_rows(min_row=2):
            if row[0].value == index:
                row[1].value = objective
                row[2].value = date
                row[3].value = person
                row[4].value = expectation
                row[5].value = results
                row[6].value = remark
                row[7].value = status
                break
        self.wb.save(self.filename)
    
    def complete_test_case(self, index, results, remark):
        for row in self.ws.iter_rows(min_row=2):
            if row[0].value == index:
                row[5].value = results
                row[6].value = remark
                row[7].value = "Completed"
                break
        self.wb.save(self.filename)
    
    def search_test_cases(self, keyword):
        results = []
        for row in self.ws.iter_rows(min_row=2):
            if any(str(cell.value).lower().find(keyword.lower()) != -1 for cell in row):
                results.append([cell.value for cell in row])
        return results
    
    def delete_test_case(self, index):
        for row in range(2, self.ws.max_row + 1):
            if self.ws.cell(row=row, column=1).value == index:
                self.ws.delete_rows(row)
                break
        self.wb.save(self.filename)
    
    def list_all_test_cases(self):
        return [[cell.value for cell in row] for row in self.ws.iter_rows(min_row=2)]

def print_test_cases(test_cases, headers):
    if not test_cases:
        print("No test cases found")
        return
    
    # Format None values as empty strings
    formatted_cases = []
    for case in test_cases:
        formatted_cases.append(["" if value is None else value for value in case])
    
    print(tabulate(formatted_cases, headers=headers, tablefmt="grid", maxcolwidths=20))

def print_help():
    print("\nAvailable commands:")
    print("  add,<objective>,<person>,<expectation>")
    print("  update,<index>,<objective>,<date>,<person>,<expectation>,<results>,<remark>,<status>")
    print("  complete,<index>,<results>,<remark>")
    print("  search,<keyword>")
    print("  delete,<index>")
    print("  list - Show all test cases")
    print("  help - Show this help")
    print("  exit - Quit the program\n")

def main():
    tracker = TestCaseTracker()
    print("Test Case Tracker System (type 'help' for commands, 'exit' to quit)")
    
    while True:
        try:
            command = input("> ").strip()
            
            if not command:
                continue
                
            if command.lower() == 'exit':
                break
                
            if command.lower() == 'help':
                print_help()
                continue
                
            if command.lower() == 'list':
                cases = tracker.list_all_test_cases()
                print(f"\nFound {len(cases)} test cases:")
                print_test_cases(cases, tracker.headers)
                continue
                
            if command.startswith("add,"):
                parts = command.split(",", 3)
                if len(parts) != 4:
                    print("Error: add requires 3 parameters")
                    continue
                _, objective, person, expectation = parts
                case_id = tracker.add_test_case(objective, person, expectation)
                print(f"Added test case with ID: {case_id}")
                
            elif command.startswith("update,"):
                parts = command.split(",")
                if len(parts) != 9:
                    print("Error: update requires 8 parameters")
                    continue
                _, index, objective, date, person, expectation, results, remark, status = parts
                tracker.update_test_case(int(index), objective, date, person, expectation, results, remark, status)
                print(f"Updated test case {index}")
                
            elif command.startswith("complete,"):
                parts = command.split(",")
                if len(parts) != 4:
                    print("Error: complete requires 3 parameters")
                    continue
                _, index, results, remark = parts
                tracker.complete_test_case(int(index), results, remark)
                print(f"Marked test case {index} as completed")
                
            elif command.startswith("search,"):
                parts = command.split(",", 1)
                if len(parts) != 2:
                    print("Error: search requires 1 parameter")
                    continue
                _, keyword = parts
                results = tracker.search_test_cases(keyword)
                print(f"\nFound {len(results)} matching test cases for '{keyword}':")
                print_test_cases(results, tracker.headers)
                    
            elif command.startswith("delete,"):
                parts = command.split(",", 1)
                if len(parts) != 2:
                    print("Error: delete requires 1 parameter")
                    continue
                _, index = parts
                tracker.delete_test_case(int(index))
                print(f"Deleted test case {index}")
                
            else:
                print("Error: Unknown command. Type 'help' for available commands")
                
        except Exception as e:
            print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()