import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from tabulate import tabulate
import sys

class BugTracker:
    def __init__(self, filename="bug_tracker.xlsx"):
        self.filename = filename
        self.headers = ["Index", "Date", "Bug", "Description", "Solution", "Person", "Files", "Status"]
        try:
            self.wb = openpyxl.load_workbook(filename)
            self.ws = self.wb.active
        except FileNotFoundError:
            self._initialize_workbook()
    
    def _initialize_workbook(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Bug Tracker"
        headers = ["Index", "Date", "Bug", "Description", "Solution", "Person", "Files", "Status"]
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        self.wb.save(self.filename)
    
    def _get_next_index(self):
        return self.ws.max_row
    
    def add_bug(self, bug, description):
        index = self._get_next_index()
        self.ws.append([
            index,
            datetime.now().strftime("%Y-%m-%d"),
            bug,
            description,
            "",  # Solution
            "",  # Person
            "",  # Files
            "Unsolved"  # Status
        ])
        self.wb.save(self.filename)
        return index
    
    def update_bug(self, index, **kwargs):
        """Update specific fields of a bug without resetting others.
        
        Args:
            index: The bug index to update
            kwargs: Any of date, bug, description, solution, 
                   person, files, status to update
        """
        column_map = {
            'date': 2,
            'bug': 3,
            'description': 4,
            'solution': 5,
            'person': 6,
            'files': 7,
            'status': 8
        }
        
        for row in self.ws.iter_rows(min_row=2):
            if row[0].value == index:
                for field, value in kwargs.items():
                    if field in column_map:
                        col = column_map[field]
                        row[col-1].value = value if value != '' else row[col-1].value
                break
        self.wb.save(self.filename)
    
    def solved_bug(self, index, solution, person, files):
        for row in self.ws.iter_rows(min_row=2):
            if row[0].value == index:
                row[4].value = solution
                row[5].value = person
                row[6].value = files
                row[7].value = "Solved"
                break
        self.wb.save(self.filename)
    
    def search_bug(self, keyword):
        results = []
        for row in self.ws.iter_rows(min_row=2):
            if any(str(cell.value).lower().find(keyword.lower()) != -1 for cell in row):
                results.append([cell.value for cell in row])
        return results
    
    def delete_bug(self, index):
        for row in range(2, self.ws.max_row + 1):
            if self.ws.cell(row=row, column=1).value == index:
                self.ws.delete_rows(row)
                break
        self.wb.save(self.filename)
    
    def list_all_bugs(self):
        return [[cell.value for cell in row] for row in self.ws.iter_rows(min_row=2)]


def print_bugs(bugs, headers):
    if not bugs:
        print("No bugs found")
        return
    
    # Format None values as empty strings
    formatted_bugs = []
    for bug in bugs:
        formatted_bugs.append(["" if value is None else value for value in bug])
    
    print(tabulate(formatted_bugs, headers=headers, tablefmt="grid", maxcolwidths=30))
def print_help():
    print("\nAvailable commands:")
    print("  add_bug,<bug>,<description>")
    print("  update_bug,<index>,<date>,<bug>,<description>,<solution>,<person>,<files>,<status>")
    print("  solved_bug,<index>,<solution>,<person>,<files>")
    print("  search_bug,<keyword>")
    print("  delete_bug,<index>")
    print("  list - Show all bugs")
    print("  help - Show this help")
    print("  exit - Quit the program\n")

def main():
    tracker = BugTracker()
    print("Bug Tracker System (type 'help' for commands, 'exit' to quit)")
    
    while True:
        try:
            command = input("> ").strip()
            
            if not command:
                continue
                
            if command.lower() == 'exit':
                break
                
            if command.lower() == 'help':
                print_help()
                continue
                
            if command.lower() == 'list':
                bugs = tracker.list_all_bugs()
                print(f"\nFound {len(bugs)} bugs:")
                print_bugs(bugs, tracker.headers)
                continue
                
            if command.startswith("add_bug,"):
                parts = command.split(",", 2)
                if len(parts) != 3:
                    print("Error: add_bug requires 2 parameters")
                    continue
                _, bug, description = parts
                bug_id = tracker.add_bug(bug, description)
                print(f"Added bug with ID: {bug_id}")
                
            elif command.startswith("update_bug,"):
                parts = command.split(",")
                if len(parts) < 2:
                    print("Error: update_bug requires at least index")
                    continue
                index = int(parts[1])
                updates = {}
                if len(parts) > 2: updates['date'] = parts[2]
                if len(parts) > 3: updates['bug'] = parts[3]
                if len(parts) > 4: updates['description'] = parts[4]
                if len(parts) > 5: updates['solution'] = parts[5]
                if len(parts) > 6: updates['person'] = parts[6]
                if len(parts) > 7: updates['files'] = parts[7]
                if len(parts) > 8: updates['status'] = parts[8]
                tracker.update_bug(index, **updates)
                print(f"Updated bug {index}")
                
            elif command.startswith("solved_bug,"):
                parts = command.split(",")
                if len(parts) != 5:
                    print("Error: solved_bug requires 4 parameters")
                    continue
                _, index, solution, person, files = parts
                tracker.solved_bug(int(index), solution, person, files)
                print(f"Marked bug {index} as solved")
                
            elif command.startswith("search_bug,"):
                parts = command.split(",", 1)
                if len(parts) != 2:
                    print("Error: search_bug requires 1 parameter")
                    continue
                _, keyword = parts
                results = tracker.search_bug(keyword)
                print(f"\nFound {len(results)} matching bugs for '{keyword}':")
                print_bugs(results, tracker.headers)
                    
            elif command.startswith("delete_bug,"):
                parts = command.split(",", 1)
                if len(parts) != 2:
                    print("Error: delete_bug requires 1 parameter")
                    continue
                _, index = parts
                tracker.delete_bug(int(index))
                print(f"Deleted bug {index}")

            else:
                print("Error: Unknown command. Type 'help' for available commands")
                
        except Exception as e:
            print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()
